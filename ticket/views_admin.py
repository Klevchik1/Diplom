# ticket/views_admin.py
import json
import csv
from datetime import datetime, timedelta
from io import BytesIO, StringIO
from decimal import Decimal

from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group, Permission
from django.core.paginator import Paginator
from django.db.models import Count, Sum, Q, Avg, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from django.conf import settings
from .report_utils import ReportGenerator
from .pdf_utils import generate_pdf_report
from django.utils import timezone
from datetime import datetime, timedelta

from .models import (
    User, HallType, AgeRating, Genre, Country, Director, Actor,
    OperationLog, APIToken, TicketStatus,
    ImportCache, Payment, TicketGroup, Ticket, Screening, Movie, Hall,
    ActionType, ModuleType, PasswordResetRequest, EmailChangeRequest,
    Seat  # <-- ДОБАВЬТЕ ЭТУ СТРОКУ
)
from .logging_utils import OperationLogger

User = get_user_model()


def is_superuser(user):
    """Проверка, является ли пользователь суперпользователем"""
    return user.is_authenticated and user.is_superuser


@staff_member_required
def admin_dashboard(request):
    """Главная страница админ-панели"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    # Статистика
    total_users = User.objects.count()
    active_users_today = User.objects.filter(last_login__date=timezone.now().date()).count()
    verified_users = User.objects.filter(is_email_verified=True).count()

    total_movies = Movie.objects.count()
    total_screenings = Screening.objects.count()
    total_tickets = Ticket.objects.count()
    total_revenue = Ticket.objects.filter(status__code='active').aggregate(
        total=Sum('price')
    )['total'] or 0

    # Логи за последние 7 дней
    week_ago = timezone.now() - timedelta(days=7)
    logs_last_week = OperationLog.objects.filter(timestamp__gte=week_ago).count()

    # Платежи
    total_payments = Payment.objects.count()
    successful_payments = Payment.objects.filter(status='succeeded').count()

    context = {
        'total_users': total_users,
        'active_users_today': active_users_today,
        'verified_users': verified_users,
        'total_movies': total_movies,
        'total_screenings': total_screenings,
        'total_tickets': total_tickets,
        'total_revenue': total_revenue,
        'logs_last_week': logs_last_week,
        'total_payments': total_payments,
        'successful_payments': successful_payments,
        'now': timezone.now(),
        'recent_logs': OperationLog.objects.select_related('user', 'action_type').all().order_by('-timestamp')[:20],
        'total_halls': Hall.objects.count(),
    }

    OperationLogger.log_operation(
        request=request,
        action_type='VIEW',
        module_type='SYSTEM',
        description=f'Просмотр дашборда админ-панели',
        additional_data={'section': 'dashboard'}
    )

    return render(request, 'ticket/admin_panel/dashboard.html', context)


@staff_member_required
def admin_users(request):
    """Управление пользователями"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    # Фильтры
    search = request.GET.get('search', '')
    is_verified = request.GET.get('is_verified', '')
    is_active = request.GET.get('is_active', '')
    is_staff = request.GET.get('is_staff', '')

    users = User.objects.all().order_by('-date_joined')

    if search:
        users = users.filter(
            Q(email__icontains=search) |
            Q(name__icontains=search) |
            Q(surname__icontains=search) |
            Q(number__icontains=search)
        )

    if is_verified == 'yes':
        users = users.filter(is_email_verified=True)
    elif is_verified == 'no':
        users = users.filter(is_email_verified=False)

    if is_active == 'yes':
        users = users.filter(is_active=True)
    elif is_active == 'no':
        users = users.filter(is_active=False)

    if is_staff == 'yes':
        users = users.filter(is_staff=True)
    elif is_staff == 'no':
        users = users.filter(is_staff=False)

    # Пагинация
    paginator = Paginator(users, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/users.html', {
        'users': page_obj,
        'search': search,
        'filters': {
            'is_verified': is_verified,
            'is_active': is_active,
            'is_staff': is_staff,
        }
    })


@staff_member_required
def admin_user_edit(request, user_id):
    """Редактирование пользователя"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        # Обновление данных
        user.name = request.POST.get('name', user.name)
        user.surname = request.POST.get('surname', user.surname)
        user.number = request.POST.get('number', user.number)
        user.is_active = request.POST.get('is_active') == 'on'
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.is_email_verified = request.POST.get('is_email_verified') == 'on'

        # Группы
        group_ids = request.POST.getlist('groups')
        user.groups.set(group_ids)

        user.save()

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='USERS',
            description=f'Администратор обновил пользователя: {user.email}',
            object_id=user.id,
            object_repr=str(user),
            additional_data={
                'changed_fields': ['name', 'surname', 'number', 'is_active', 'is_staff', 'is_email_verified']
            }
        )

        messages.success(request, f'Пользователь {user.email} успешно обновлён.')
        return redirect('admin_panel_users')

    groups = Group.objects.all()
    user_groups = user.groups.values_list('id', flat=True)

    return render(request, 'ticket/admin_panel/user_edit.html', {
        'edit_user': user,
        'groups': groups,
        'user_groups': user_groups,
    })


@staff_member_required
@require_POST
def admin_user_delete(request, user_id):
    """Удаление пользователя"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.error(request, 'Нельзя удалить самого себя.')
        return redirect('admin_panel_users')

    email = user.email

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='USERS',
        description=f'Администратор удалил пользователя: {email}',
        object_id=user.id,
        object_repr=str(user)
    )

    user.delete()
    messages.success(request, f'Пользователь {email} удалён.')
    return redirect('admin_panel_users')


@staff_member_required
@require_POST
def admin_user_toggle_block(request, user_id):
    """Блокировка/разблокировка пользователя"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.error(request, 'Нельзя заблокировать самого себя.')
        return redirect('admin_panel_users')

    user.is_active = not user.is_active
    user.save()

    status = 'заблокирован' if not user.is_active else 'разблокирован'

    OperationLogger.log_operation(
        request=request,
        action_type='UPDATE',
        module_type='USERS',
        description=f'Администратор {status} пользователя: {user.email}',
        object_id=user.id,
        object_repr=str(user)
    )

    messages.success(request, f'Пользователь {user.email} {status}.')
    return redirect('admin_panel_users')


@staff_member_required
def admin_roles(request):
    """Управление ролями (группами)"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    groups = Group.objects.all().annotate(
        user_count=Count('user')
    ).order_by('name')

    return render(request, 'ticket/admin_panel/roles.html', {'groups': groups})


@staff_member_required
def admin_role_edit(request, group_id):
    """Редактирование роли"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    group = get_object_or_404(Group, id=group_id)

    if request.method == 'POST':
        group.name = request.POST.get('name', group.name)

        # Обновление разрешений
        permission_ids = request.POST.getlist('permissions')
        group.permissions.set(permission_ids)
        group.save()

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='SYSTEM',
            description=f'Администратор обновил роль: {group.name}',
            object_id=group.id,
            object_repr=group.name
        )

        messages.success(request, f'Роль "{group.name}" успешно обновлена.')
        return redirect('admin_panel_roles')

    all_permissions = Permission.objects.select_related('content_type').order_by('content_type__app_label', 'codename')
    group_permissions = group.permissions.values_list('id', flat=True)

    # Группировка разрешений по приложениям
    permissions_by_app = {}
    for perm in all_permissions:
        app_label = perm.content_type.app_label
        if app_label not in permissions_by_app:
            permissions_by_app[app_label] = []
        permissions_by_app[app_label].append({
            'id': perm.id,
            'name': perm.name,
            'codename': perm.codename,
            'checked': perm.id in group_permissions
        })

    return render(request, 'ticket/admin_panel/role_edit.html', {
        'group': group,
        'permissions_by_app': permissions_by_app,
    })


@staff_member_required
def admin_hall_types(request):
    """Управление типами залов"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    hall_types = HallType.objects.all().order_by('name')

    if search:
        hall_types = hall_types.filter(
            Q(name__icontains=search) | Q(description__icontains=search)
        )

    # Пагинация
    paginator = Paginator(hall_types, 20)
    page_number = request.GET.get('page', 1)
    hall_types_page = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        base_price = request.POST.get('base_price')
        price_coefficient = request.POST.get('price_coefficient', '1.0')
        description = request.POST.get('description', '')

        # Валидация
        errors = []

        if not name:
            errors.append('Название обязательно для заполнения')
        if not base_price:
            errors.append('Базовая цена обязательна для заполнения')

        # Проверка коэффициента
        try:
            coeff = float(price_coefficient)
            if coeff <= 0:
                errors.append('Коэффициент должен быть больше 0')
            elif coeff > 9.99:
                errors.append('⚠️ Коэффициент не может быть больше 9.99! Максимальное значение: 9.99')
        except ValueError:
            errors.append('Коэффициент должен быть числом')

        # Проверка базовой цены
        try:
            price = float(base_price)
            if price <= 0:
                errors.append('Базовая цена должна быть больше 0')
            elif price > 999999:
                errors.append('Базовая цена не может быть больше 999 999 ₽')
        except ValueError:
            errors.append('Базовая цена должна быть числом')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                from decimal import Decimal
                hall_type = HallType.objects.create(
                    name=name,
                    base_price=Decimal(str(base_price)),
                    price_coefficient=Decimal(str(price_coefficient)),
                    description=description
                )
                OperationLogger.log_operation(
                    request=request,
                    action_type='CREATE',
                    module_type='SYSTEM',
                    description=f'Создан тип зала: {name}',
                    object_id=hall_type.id,
                    object_repr=name
                )
                messages.success(request, f'Тип зала "{name}" успешно создан.')
                return redirect('admin_panel_hall_types')
            except Exception as e:
                messages.error(request, f'Ошибка при создании: {str(e)}')

    return render(request, 'ticket/admin_panel/hall_types.html', {
        'hall_types': hall_types_page,
        'search': search,
    })


@staff_member_required
def admin_hall_type_edit(request, ht_id):
    """Редактирование типа зала"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    hall_type = get_object_or_404(HallType, id=ht_id)

    if request.method == 'POST':
        name = request.POST.get('name', hall_type.name)
        base_price = request.POST.get('base_price', str(hall_type.base_price))
        price_coefficient = request.POST.get('price_coefficient', str(hall_type.price_coefficient))
        description = request.POST.get('description', hall_type.description)

        # Валидация
        errors = []

        if not name:
            errors.append('Название обязательно для заполнения')
        if not base_price:
            errors.append('Базовая цена обязательна для заполнения')

        # Проверка коэффициента
        try:
            coeff = float(price_coefficient)
            if coeff <= 0:
                errors.append('Коэффициент должен быть больше 0')
            elif coeff > 9.99:
                errors.append('⚠️ Коэффициент не может быть больше 9.99! Максимальное значение: 9.99')
        except ValueError:
            errors.append('Коэффициент должен быть числом')

        # Проверка базовой цены
        try:
            price = float(base_price)
            if price <= 0:
                errors.append('Базовая цена должна быть больше 0')
            elif price > 999999:
                errors.append('Базовая цена не может быть больше 999 999 ₽')
        except ValueError:
            errors.append('Базовая цена должна быть числом')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                from decimal import Decimal
                old_name = hall_type.name
                hall_type.name = name
                hall_type.base_price = Decimal(str(base_price))
                hall_type.price_coefficient = Decimal(str(price_coefficient))
                hall_type.description = description
                hall_type.save()

                OperationLogger.log_operation(
                    request=request,
                    action_type='UPDATE',
                    module_type='SYSTEM',
                    description=f'Обновлён тип зала: {old_name} → {hall_type.name}',
                    object_id=hall_type.id,
                    object_repr=hall_type.name
                )

                messages.success(request, f'Тип зала "{hall_type.name}" успешно обновлён.')
                return redirect('admin_panel_hall_types')
            except Exception as e:
                messages.error(request, f'Ошибка при обновлении: {str(e)}')

    return render(request, 'ticket/admin_panel/hall_type_edit.html', {'hall_type': hall_type})


@staff_member_required
def admin_hall_type_delete(request, ht_id):
    """Удаление типа зала"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    hall_type = get_object_or_404(HallType, id=ht_id)

    if request.method == 'POST':
        name = hall_type.name
        hall_type.delete()

        OperationLogger.log_operation(
            request=request,
            action_type='DELETE',
            module_type='SYSTEM',
            description=f'Удалён тип зала: {name}',
            object_id=ht_id,
            object_repr=name
        )

        messages.success(request, f'Тип зала "{name}" удалён.')
        return redirect('admin_panel_hall_types')

    return render(request, 'ticket/admin_panel/hall_type_confirm_delete.html', {'hall_type': hall_type})


@staff_member_required
def admin_age_ratings(request):
    """Управление возрастными рейтингами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    age_ratings = AgeRating.objects.all().order_by('name')

    if search:
        age_ratings = age_ratings.filter(name__icontains=search)

    paginator = Paginator(age_ratings, 20)
    page_number = request.GET.get('page', 1)
    age_ratings_page = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            age_rating = AgeRating.objects.create(name=name)
            OperationLogger.log_operation(
                request=request,
                action_type='CREATE',
                module_type='SYSTEM',
                description=f'Создан возрастной рейтинг: {name}',
                object_id=age_rating.id,
                object_repr=name
            )
            messages.success(request, f'Возрастной рейтинг "{name}" создан.')
            return redirect('admin_panel_age_ratings')

    return render(request, 'ticket/admin_panel/age_ratings.html', {
        'age_ratings': age_ratings_page,
        'search': search,
    })


@staff_member_required
@require_POST
def admin_age_rating_delete(request, ar_id):
    """Удаление возрастного рейтинга"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    age_rating = get_object_or_404(AgeRating, id=ar_id)
    name = age_rating.name
    age_rating.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён возрастной рейтинг: {name}',
        object_id=ar_id,
        object_repr=name
    )

    messages.success(request, f'Возрастной рейтинг "{name}" удалён.')
    return redirect('admin_panel_age_ratings')


@staff_member_required
def admin_genres(request):
    """Управление жанрами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    genres = Genre.objects.all().order_by('name')

    if search:
        genres = genres.filter(
            Q(name__icontains=search) | Q(description__icontains=search)
        )

    paginator = Paginator(genres, 20)
    page_number = request.GET.get('page', 1)
    genres_page = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        if name:
            genre = Genre.objects.create(name=name, description=description)
            OperationLogger.log_operation(
                request=request,
                action_type='CREATE',
                module_type='SYSTEM',
                description=f'Создан жанр: {name}',
                object_id=genre.id,
                object_repr=name
            )
            messages.success(request, f'Жанр "{name}" создан.')
            return redirect('admin_panel_genres')

    return render(request, 'ticket/admin_panel/genres.html', {
        'genres': genres_page,
        'search': search,
    })


@staff_member_required
@require_POST
def admin_genre_delete(request, genre_id):
    """Удаление жанра"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    genre = get_object_or_404(Genre, id=genre_id)
    name = genre.name
    genre.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён жанр: {name}',
        object_id=genre_id,
        object_repr=name
    )

    messages.success(request, f'Жанр "{name}" удалён.')
    return redirect('admin_panel_genres')


@staff_member_required
def admin_countries(request):
    """Управление странами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    countries = Country.objects.all().order_by('name')

    if search:
        countries = countries.filter(
            Q(name__icontains=search) | Q(code__icontains=search)
        )

    paginator = Paginator(countries, 20)
    page_number = request.GET.get('page', 1)
    countries_page = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code', '').upper()
        if name and code:
            country = Country.objects.create(name=name, code=code)
            OperationLogger.log_operation(
                request=request,
                action_type='CREATE',
                module_type='SYSTEM',
                description=f'Создана страна: {name} ({code})',
                object_id=country.id,
                object_repr=name
            )
            messages.success(request, f'Страна "{name}" создана.')
            return redirect('admin_panel_countries')

    return render(request, 'ticket/admin_panel/countries.html', {
        'countries': countries_page,
        'search': search,
    })


@staff_member_required
@require_POST
def admin_country_delete(request, country_id):
    """Удаление страны"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    country = get_object_or_404(Country, id=country_id)
    name = country.name
    country.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалена страна: {name}',
        object_id=country_id,
        object_repr=name
    )

    messages.success(request, f'Страна "{name}" удалена.')
    return redirect('admin_panel_countries')


@staff_member_required
def admin_directors(request):
    """Управление режиссёрами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    country_id = request.GET.get('country', '')

    directors = Director.objects.all().order_by('surname', 'name')

    if search:
        directors = directors.filter(
            Q(name__icontains=search) | Q(surname__icontains=search)
        )
    if country_id:
        directors = directors.filter(country_id=country_id)

    paginator = Paginator(directors, 20)
    page_number = request.GET.get('page', 1)
    directors_page = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        surname = request.POST.get('surname')
        birth_date_str = request.POST.get('birth_date', '').strip()
        country_id_val = request.POST.get('country')
        biography = request.POST.get('biography', '')

        # Валидация
        errors = []

        if not name:
            errors.append('Имя обязательно для заполнения')
        if not surname:
            errors.append('Фамилия обязательна для заполнения')

        # Валидация даты рождения
        if birth_date_str:
            from datetime import date
            try:
                birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                today = date.today()

                # Проверка: дата не в будущем
                if birth_date > today:
                    errors.append('Дата рождения не может быть в будущем')

                # Проверка: не раньше 1800 года
                if birth_date.year < 1800:
                    errors.append('Дата рождения не может быть раньше 1800 года')

                # Проверка: возраст не менее 18 лет (если дата указана)
                age = today.year - birth_date.year
                if (today.month, today.day) < (birth_date.month, birth_date.day):
                    age -= 1

                if age < 18:
                    errors.append(
                        f'⚠️ Режиссёру должно быть не менее 18 лет. Текущий возраст: {age} лет. Минимальная дата рождения: {today.year - 18}-{today.month:02d}-{today.day:02d}')

            except ValueError:
                errors.append('Неверный формат даты рождения')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                director = Director.objects.create(
                    name=name,
                    surname=surname,
                    birth_date=birth_date if birth_date_str else None,
                    country_id=country_id_val if country_id_val else None,
                    biography=biography
                )
                OperationLogger.log_operation(
                    request=request,
                    action_type='CREATE',
                    module_type='SYSTEM',
                    description=f'Создан режиссёр: {name} {surname}',
                    object_id=director.id,
                    object_repr=str(director)
                )
                messages.success(request, f'Режиссёр "{name} {surname}" создан.')
                return redirect('admin_panel_directors')
            except Exception as e:
                messages.error(request, f'Ошибка при создании: {str(e)}')

    countries = Country.objects.all().order_by('name')

    return render(request, 'ticket/admin_panel/directors.html', {
        'directors': directors_page,
        'countries': countries,
        'search': search,
    })


@require_POST
def admin_director_delete(request, director_id):
    """Удаление режиссёра"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    director = get_object_or_404(Director, id=director_id)
    name = f"{director.name} {director.surname}"
    director.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён режиссёр: {name}',
        object_id=director_id,
        object_repr=name
    )

    messages.success(request, f'Режиссёр "{name}" удалён.')
    return redirect('admin_panel_directors')

@staff_member_required
def admin_actors(request):
    """Управление актёрами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    country_id = request.GET.get('country', '')

    actors = Actor.objects.all().order_by('surname', 'name')

    if search:
        actors = actors.filter(
            Q(name__icontains=search) | Q(surname__icontains=search)
        )
    if country_id:
        actors = actors.filter(country_id=country_id)

    paginator = Paginator(actors, 20)
    page_number = request.GET.get('page', 1)
    actors_page = paginator.get_page(page_number)

    if request.method == 'POST':
        name = request.POST.get('name')
        surname = request.POST.get('surname')
        birth_date_str = request.POST.get('birth_date', '').strip()
        country_id_val = request.POST.get('country')
        biography = request.POST.get('biography', '')

        # Валидация
        errors = []

        if not name:
            errors.append('Имя обязательно для заполнения')
        if not surname:
            errors.append('Фамилия обязательна для заполнения')

        # Валидация даты рождения
        if birth_date_str:
            from datetime import date
            try:
                birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                today = date.today()

                # Проверка: дата не в будущем
                if birth_date > today:
                    errors.append('Дата рождения не может быть в будущем')

                # Проверка: не раньше 1800 года
                if birth_date.year < 1800:
                    errors.append('Дата рождения не может быть раньше 1800 года')

                # Проверка: возраст не менее 18 лет
                age = today.year - birth_date.year
                if (today.month, today.day) < (birth_date.month, birth_date.day):
                    age -= 1

                if age < 18:
                    min_birth_date = date(today.year - 18, today.month, today.day)
                    errors.append(
                        f'⚠️ Актёру должно быть не менее 18 лет. Текущий возраст: {age} лет. Минимальная дата рождения: {min_birth_date.strftime("%d.%m.%Y")}')

            except ValueError:
                errors.append('Неверный формат даты рождения')

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                actor = Actor.objects.create(
                    name=name,
                    surname=surname,
                    birth_date=birth_date if birth_date_str else None,
                    country_id=country_id_val if country_id_val else None,
                    biography=biography
                )
                OperationLogger.log_operation(
                    request=request,
                    action_type='CREATE',
                    module_type='SYSTEM',
                    description=f'Создан актёр: {name} {surname}',
                    object_id=actor.id,
                    object_repr=str(actor)
                )
                messages.success(request, f'Актёр "{name} {surname}" создан.')
                return redirect('admin_panel_actors')
            except Exception as e:
                messages.error(request, f'Ошибка при создании: {str(e)}')

    countries = Country.objects.all().order_by('name')

    return render(request, 'ticket/admin_panel/actors.html', {
        'actors': actors_page,
        'countries': countries,
        'search': search,
    })


@require_POST
def admin_actor_delete(request, actor_id):
    """Удаление актёра"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    actor = get_object_or_404(Actor, id=actor_id)
    name = f"{actor.name} {actor.surname}"
    actor.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён актёр: {name}',
        object_id=actor_id,
        object_repr=name
    )

    messages.success(request, f'Актёр "{name}" удалён.')
    return redirect('admin_panel_actors')


@staff_member_required
def admin_logs(request):
    """Просмотр логов операций"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    # Фильтры
    action_type = request.GET.get('action_type', '')
    module_type = request.GET.get('module_type', '')
    user_id = request.GET.get('user_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    logs = OperationLog.objects.select_related('user', 'action_type', 'module_type').all()

    if action_type:
        logs = logs.filter(action_type__code=action_type)
    if module_type:
        logs = logs.filter(module_type__code=module_type)
    if user_id:
        logs = logs.filter(user_id=user_id)
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)

    logs = logs.order_by('-timestamp')

    paginator = Paginator(logs, 100)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Для фильтров - используем правильные имена моделей
    action_types = ActionType.objects.all().order_by('name')
    module_types = ModuleType.objects.all().order_by('name')
    users = User.objects.all().order_by('email')

    return render(request, 'ticket/admin_panel/logs.html', {
        'logs': page_obj,
        'action_types': action_types,
        'module_types': module_types,
        'users': users,
        'filters': {
            'action_type': action_type,
            'module_type': module_type,
            'user_id': user_id,
            'date_from': date_from,
            'date_to': date_to,
        }
    })


@staff_member_required
def admin_logs_export(request):
    """Экспорт логов в CSV/JSON/PDF/XLSX"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    format_type = request.GET.get('format', 'csv')

    # Получаем параметры фильтрации
    action_type = request.GET.get('action_type', '')
    module_type = request.GET.get('module_type', '')
    user_id = request.GET.get('user_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # Строим запрос
    logs = OperationLog.objects.select_related('user', 'action_type', 'module_type').all()

    if action_type:
        logs = logs.filter(action_type__code=action_type)
    if module_type:
        logs = logs.filter(module_type__code=module_type)
    if user_id and user_id.isdigit():
        logs = logs.filter(user_id=int(user_id))
    if date_from:
        try:
            logs = logs.filter(timestamp__date__gte=date_from)
        except:
            pass
    if date_to:
        try:
            logs = logs.filter(timestamp__date__lte=date_to)
        except:
            pass

    logs = logs.order_by('-timestamp')

    # Логируем экспорт
    OperationLogger.log_operation(
        request=request,
        action_type='EXPORT',
        module_type='SYSTEM',
        description=f'Экспорт логов в {format_type.upper()} ({logs.count()} записей)'
    )

    if format_type == 'csv':
        return export_logs_csv(logs)
    elif format_type == 'xlsx':
        return export_logs_xlsx(logs)
    elif format_type == 'json':
        return export_logs_json(logs)
    elif format_type == 'pdf':
        return export_logs_pdf(request, logs)
    else:
        messages.error(request, 'Неподдерживаемый формат экспорта.')
        return redirect('admin_panel_logs')


def export_logs_csv(logs):
    """Экспорт в CSV с правильным разделением для Excel"""
    import csv

    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="logs_export.csv"'

    writer = csv.writer(response, delimiter=';')  # Используем ; как разделитель для Excel

    # Заголовки
    writer.writerow([
        'ID',
        'Пользователь',
        'Тип действия',
        'Модуль',
        'Описание',
        'IP адрес',
        'User Agent',
        'Время',
        'Object ID',
        'Object Repr'
    ])

    # Данные
    for log in logs:
        writer.writerow([
            log.id,
            log.user.email if log.user else 'Аноним',
            log.action_type.name if log.action_type else '-',
            log.module_type.name if log.module_type else '-',
            log.description,
            log.ip_address or '-',
            (log.user_agent or '-')[:100],  # Ограничиваем длину
            log.timestamp.strftime('%d.%m.%Y %H:%M:%S'),
            log.object_id or '-',
            log.object_repr or '-',
        ])

    return response


def export_logs_xlsx(logs):
    """Экспорт в Excel XLSX с автошириной колонок"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Логи операций"

    # Заголовки
    headers = ['ID', 'Пользователь', 'Тип действия', 'Модуль', 'Описание', 'IP адрес', 'User Agent', 'Время',
               'Object ID', 'Object Repr']
    ws.append(headers)

    # Стиль заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Данные
    for log in logs:
        ws.append([
            log.id,
            log.user.email if log.user else 'Аноним',
            log.action_type.name if log.action_type else '-',
            log.module_type.name if log.module_type else '-',
            log.description,
            log.ip_address or '-',
            (log.user_agent or '-')[:100],
            log.timestamp.strftime('%d.%m.%Y %H:%M:%S'),
            log.object_id or '-',
            log.object_repr or '-',
        ])

    # Автоподбор ширины колонок
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = len(headers[col - 1])
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="logs_export.xlsx"'
    wb.save(response)
    return response


def export_logs_json(logs):
    """Экспорт в JSON"""
    import json

    data = []
    for log in logs:
        data.append({
            'id': log.id,
            'user': log.user.email if log.user else None,
            'action': log.action_type.name if log.action_type else None,
            'module': log.module_type.name if log.module_type else None,
            'description': log.description,
            'ip_address': log.ip_address,
            'timestamp': log.timestamp.isoformat(),
            'object_id': log.object_id,
            'object_repr': log.object_repr,
            'additional_data': log.additional_data,
        })

    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = 'attachment; filename="logs_export.json"'
    return response


def export_logs_pdf(request, logs):
    """Экспорт в PDF с шапкой документа"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import mm
    import os
    from django.conf import settings
    from datetime import datetime

    # Регистрируем шрифт для кириллицы
    font_path = os.path.join(settings.BASE_DIR, 'ticket', 'fonts', 'DejaVuSans.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
        font_name = 'DejaVuSans'
    else:
        font_name = 'Helvetica'

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="logs_export.pdf"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4),
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=20 * mm, bottomMargin=15 * mm)

    # Получаем базовые стили
    styles = getSampleStyleSheet()

    # Создаём новые стили с уникальными именами
    russian_style = ParagraphStyle(
        'RussianStyle', parent=styles['Normal'],
        fontName=font_name, fontSize=9, leading=12
    )
    russian_bold = ParagraphStyle(
        'RussianBold', parent=styles['Normal'],
        fontName=font_name, fontSize=9, leading=12,
        fontWeight='bold'
    )
    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'],
        fontName=font_name, fontSize=16, leading=24,
        alignment=1, fontWeight='bold'
    )
    header_style = ParagraphStyle(
        'ReportHeader', parent=styles['Heading2'],
        fontName=font_name, fontSize=12, leading=16,
        fontWeight='bold'
    )
    info_style = ParagraphStyle(
        'ReportInfo', parent=styles['Normal'],
        fontName=font_name, fontSize=10, leading=14
    )

    elements = []

    # ========== ШАПКА ДОКУМЕНТА ==========

    # Заголовок
    elements.append(Paragraph("КИНОТЕАТР «ПРЕМЬЕРА»", title_style))
    elements.append(Spacer(1, 5))
    elements.append(Paragraph("Отчёт по логам операций", header_style))
    elements.append(Spacer(1, 10))

    # Дата формирования
    elements.append(Paragraph(
        f"<b>Дата формирования отчёта:</b> {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}",
        info_style
    ))

    # Пользователь, сформировавший отчёт
    user_name = f"{request.user.name} {request.user.surname}" if request.user.name else request.user.email
    elements.append(Paragraph(
        f"<b>Сформировал:</b> {user_name}",
        info_style
    ))
    elements.append(Spacer(1, 5))

    # Информация о фильтрах
    elements.append(Paragraph("<b>Параметры фильтрации:</b>", info_style))

    # Получаем параметры фильтрации
    action_type = request.GET.get('action_type', '')
    module_type = request.GET.get('module_type', '')
    user_id = request.GET.get('user_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    filters_text = []
    if action_type:
        try:
            from .models import ActionType
            at = ActionType.objects.get(code=action_type)
            filters_text.append(f"Тип действия: {at.name}")
        except:
            filters_text.append(f"Тип действия: {action_type}")
    if module_type:
        try:
            from .models import ModuleType
            mt = ModuleType.objects.get(code=module_type)
            filters_text.append(f"Модуль: {mt.name}")
        except:
            filters_text.append(f"Модуль: {module_type}")
    if user_id:
        try:
            from .models import User
            u = User.objects.get(id=user_id)
            filters_text.append(f"Пользователь: {u.email}")
        except:
            filters_text.append(f"Пользователь ID: {user_id}")
    if date_from:
        filters_text.append(f"Дата от: {date_from}")
    if date_to:
        filters_text.append(f"Дата до: {date_to}")

    if filters_text:
        for ft in filters_text:
            elements.append(Paragraph(f"&nbsp;&nbsp;• {ft}", info_style))
    else:
        elements.append(Paragraph("&nbsp;&nbsp;• Без фильтров (все записи)", info_style))

    elements.append(Spacer(1, 5))

    # Количество записей
    total_count = logs.count()
    displayed_count = min(total_count, 500)
    elements.append(Paragraph(
        f"<b>Количество записей в отчёте:</b> {displayed_count} из {total_count}" +
        (" (отображены первые 500)" if total_count > 500 else ""),
        info_style
    ))
    elements.append(Spacer(1, 15))

    # ========== ТАБЛИЦА ==========

    # Заголовки таблицы
    table_data = [[
        Paragraph("<b>ID</b>", russian_style),
        Paragraph("<b>Пользователь</b>", russian_style),
        Paragraph("<b>Действие</b>", russian_style),
        Paragraph("<b>Модуль</b>", russian_style),
        Paragraph("<b>Описание</b>", russian_style),
        Paragraph("<b>IP</b>", russian_style),
        Paragraph("<b>Время</b>", russian_style),
    ]]

    # Данные (ограничиваем 500 для производительности)
    for log in logs[:500]:
        table_data.append([
            Paragraph(str(log.id), russian_style),
            Paragraph(log.user.email[:25] if log.user else 'Аноним', russian_style),
            Paragraph(log.action_type.name[:20] if log.action_type else '-', russian_style),
            Paragraph(log.module_type.name[:20] if log.module_type else '-', russian_style),
            Paragraph(log.description[:60] + ('...' if len(log.description) > 60 else ''), russian_style),
            Paragraph(log.ip_address or '-', russian_style),
            Paragraph(log.timestamp.strftime('%d.%m.%Y %H:%M'), russian_style),
        ])

    # Создаём таблицу с автоматической шириной колонок
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#bdc3c7')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8f9fa')]),
    ]))

    elements.append(table)

    # ========== ПОДВАЛ ==========
    elements.append(Spacer(1, 15))
    elements.append(Paragraph(
        f"<i>Отчёт сгенерирован автоматически. Всего записей: {total_count}.</i>",
        russian_style
    ))

    # Строим документ
    doc.build(elements)

    return response


@staff_member_required
def admin_ticket_statuses(request):
    """Управление статусами билетов"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    statuses = TicketStatus.objects.all().order_by('id')

    if request.method == 'POST':
        status_id = request.POST.get('status_id')
        status = get_object_or_404(TicketStatus, id=status_id)

        old_name = status.name
        status.name = request.POST.get('name', status.name)
        status.description = request.POST.get('description', status.description)
        status.is_active = request.POST.get('is_active') == 'on'
        status.can_be_refunded = request.POST.get('can_be_refunded') == 'on'
        status.save()

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='SYSTEM',
            description=f'Обновлён статус билета: {old_name} → {status.name}',
            object_id=status.id,
            object_repr=status.name
        )

        messages.warning(request,
                         f'⚠️ Статус билета "{status.name}" обновлён. Будьте осторожны: изменение статусов может повлиять на логику работы системы!')
        return redirect('admin_panel_ticket_statuses')

    return render(request, 'ticket/admin_panel/ticket_statuses.html', {'statuses': statuses})


@staff_member_required
def admin_payments(request):
    """Просмотр платежей (только чтение)"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    payments = Payment.objects.select_related('ticket_group', 'ticket_group__user',
                                              'ticket_group__screening__movie').all().order_by('-created_at')

    # Фильтры
    status = request.GET.get('status', '')
    if status:
        payments = payments.filter(status=status)

    paginator = Paginator(payments, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/payments.html', {
        'payments': page_obj,
        'filters': {'status': status}
    })


@staff_member_required
def admin_system_info(request):
    """Информация о системе"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    # Статистика БД
    db_stats = {
        'users': User.objects.count(),
        'movies': Movie.objects.count(),
        'screenings': Screening.objects.count(),
        'halls': Hall.objects.count(),
        'tickets': Ticket.objects.count(),
        'payments': Payment.objects.count(),
        'logs': OperationLog.objects.count(),
    }

    # Запросы по типам
    from .models import ActionType
    actions_stats = {}
    for action in ActionType.objects.all():
        actions_stats[action.name] = OperationLog.objects.filter(action_type=action).count()

    context = {
        'db_stats': db_stats,
        'actions_stats': actions_stats,
    }

    return render(request, 'ticket/admin_panel/system_info.html', context)


@staff_member_required
def admin_redirect_to_django(request):
    """Редирект на Django Admin"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    OperationLogger.log_operation(
        request=request,
        action_type='VIEW',
        module_type='SYSTEM',
        description=f'Переход в Django Admin из админ-панели'
    )

    return redirect('/admin/')


# ==================== ТИПЫ ДЕЙСТВИЙ (ActionType) ====================

@staff_member_required
def admin_action_types(request):
    """Управление типами действий"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    action_types = ActionType.objects.all().order_by('code')

    if search:
        action_types = action_types.filter(
            Q(code__icontains=search) | Q(name__icontains=search) | Q(description__icontains=search)
        )

    paginator = Paginator(action_types, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/action_types.html', {
        'action_types': page_obj,
        'search': search,
    })


@staff_member_required
def admin_action_type_add(request):
    """Добавление типа действия"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    if request.method == 'POST':
        code = request.POST.get('code', '').upper()
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        if ActionType.objects.filter(code=code).exists():
            messages.error(request, f'Тип действия с кодом "{code}" уже существует')
        elif ActionType.objects.filter(name=name).exists():
            messages.error(request, f'Тип действия с названием "{name}" уже существует')
        else:
            action_type = ActionType.objects.create(
                code=code,
                name=name,
                description=description
            )
            OperationLogger.log_operation(
                request=request,
                action_type='CREATE',
                module_type='SYSTEM',
                description=f'Создан тип действия: {name} ({code})',
                object_id=action_type.id,
                object_repr=name
            )
            messages.success(request, f'Тип действия "{name}" успешно создан')
            return redirect('admin_panel_action_types')

    return render(request, 'ticket/admin_panel/action_type_form.html', {'action': 'add'})


@staff_member_required
def admin_action_type_edit(request, at_id):
    """Редактирование типа действия"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    action_type = get_object_or_404(ActionType, id=at_id)

    if request.method == 'POST':
        code = request.POST.get('code', '').upper()
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        # Проверка уникальности
        if ActionType.objects.exclude(id=at_id).filter(code=code).exists():
            messages.error(request, f'Тип действия с кодом "{code}" уже существует')
        elif ActionType.objects.exclude(id=at_id).filter(name=name).exists():
            messages.error(request, f'Тип действия с названием "{name}" уже существует')
        else:
            old_name = action_type.name
            action_type.code = code
            action_type.name = name
            action_type.description = description
            action_type.save()

            OperationLogger.log_operation(
                request=request,
                action_type='UPDATE',
                module_type='SYSTEM',
                description=f'Обновлён тип действия: {old_name} → {name}',
                object_id=action_type.id,
                object_repr=name
            )
            messages.success(request, f'Тип действия "{name}" успешно обновлён')
            return redirect('admin_panel_action_types')

    return render(request, 'ticket/admin_panel/action_type_form.html', {
        'action_type': action_type,
        'action': 'edit'
    })


@require_POST
def admin_action_type_delete(request, at_id):
    """Удаление типа действия"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    action_type = get_object_or_404(ActionType, id=at_id)
    name = action_type.name

    # Проверка на связанные логи
    if OperationLog.objects.filter(action_type=action_type).exists():
        messages.error(request, f'Нельзя удалить тип действия "{name}", так как существуют связанные логи операций')
        return redirect('admin_panel_action_types')

    action_type.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён тип действия: {name}',
        object_id=at_id,
        object_repr=name
    )
    messages.success(request, f'Тип действия "{name}" удалён')
    return redirect('admin_panel_action_types')


# ==================== ТИПЫ МОДУЛЕЙ (ModuleType) ====================

@staff_member_required
def admin_module_types(request):
    """Управление типами модулей"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    module_types = ModuleType.objects.all().order_by('code')

    if search:
        module_types = module_types.filter(
            Q(code__icontains=search) | Q(name__icontains=search) | Q(description__icontains=search)
        )

    paginator = Paginator(module_types, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/module_types.html', {
        'module_types': page_obj,
        'search': search,
    })


@staff_member_required
def admin_module_type_add(request):
    """Добавление типа модуля"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    if request.method == 'POST':
        code = request.POST.get('code', '').upper()
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        if ModuleType.objects.filter(code=code).exists():
            messages.error(request, f'Тип модуля с кодом "{code}" уже существует')
        elif ModuleType.objects.filter(name=name).exists():
            messages.error(request, f'Тип модуля с названием "{name}" уже существует')
        else:
            module_type = ModuleType.objects.create(
                code=code,
                name=name,
                description=description
            )
            OperationLogger.log_operation(
                request=request,
                action_type='CREATE',
                module_type='SYSTEM',
                description=f'Создан тип модуля: {name} ({code})',
                object_id=module_type.id,
                object_repr=name
            )
            messages.success(request, f'Тип модуля "{name}" успешно создан')
            return redirect('admin_panel_module_types')

    return render(request, 'ticket/admin_panel/module_type_form.html', {'action': 'add'})


@staff_member_required
def admin_module_type_edit(request, mt_id):
    """Редактирование типа модуля"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    module_type = get_object_or_404(ModuleType, id=mt_id)

    if request.method == 'POST':
        code = request.POST.get('code', '').upper()
        name = request.POST.get('name')
        description = request.POST.get('description', '')

        if ModuleType.objects.exclude(id=mt_id).filter(code=code).exists():
            messages.error(request, f'Тип модуля с кодом "{code}" уже существует')
        elif ModuleType.objects.exclude(id=mt_id).filter(name=name).exists():
            messages.error(request, f'Тип модуля с названием "{name}" уже существует')
        else:
            old_name = module_type.name
            module_type.code = code
            module_type.name = name
            module_type.description = description
            module_type.save()

            OperationLogger.log_operation(
                request=request,
                action_type='UPDATE',
                module_type='SYSTEM',
                description=f'Обновлён тип модуля: {old_name} → {name}',
                object_id=module_type.id,
                object_repr=name
            )
            messages.success(request, f'Тип модуля "{name}" успешно обновлён')
            return redirect('admin_panel_module_types')

    return render(request, 'ticket/admin_panel/module_type_form.html', {
        'module_type': module_type,
        'action': 'edit'
    })


@require_POST
def admin_module_type_delete(request, mt_id):
    """Удаление типа модуля"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    module_type = get_object_or_404(ModuleType, id=mt_id)
    name = module_type.name

    if OperationLog.objects.filter(module_type=module_type).exists():
        messages.error(request, f'Нельзя удалить тип модуля "{name}", так как существуют связанные логи операций')
        return redirect('admin_panel_module_types')

    module_type.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён тип модуля: {name}',
        object_id=mt_id,
        object_repr=name
    )
    messages.success(request, f'Тип модуля "{name}" удалён')
    return redirect('admin_panel_module_types')


# ==================== ГРУППЫ БИЛЕТОВ (TicketGroup) ====================

@staff_member_required
def admin_ticket_groups(request):
    """Просмотр групп билетов"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')

    groups = TicketGroup.objects.select_related('user', 'screening__movie').all().order_by('-purchase_date')

    if search:
        groups = groups.filter(
            Q(group_uuid__icontains=search) |
            Q(user__email__icontains=search) |
            Q(screening__movie__title__icontains=search)
        )
    if status:
        groups = groups.filter(payment_status=status)

    paginator = Paginator(groups, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/ticket_groups.html', {
        'groups': page_obj,
        'search': search,
        'status_filter': status,
    })


@staff_member_required
def admin_ticket_group_detail(request, group_uuid):
    """Детальный просмотр группы билетов"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    group = get_object_or_404(TicketGroup, group_uuid=group_uuid)
    tickets = group.tickets.select_related('seat', 'status').all()

    return render(request, 'ticket/admin_panel/ticket_group_detail.html', {
        'group': group,
        'tickets': tickets,
    })


# ==================== МЕСТА (Seat) ====================

@staff_member_required
def admin_seats(request):
    """Просмотр мест"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    hall_id = request.GET.get('hall', '')
    search = request.GET.get('search', '')

    seats = Seat.objects.select_related('hall').all().order_by('hall__name', 'row', 'number')

    if hall_id:
        seats = seats.filter(hall_id=hall_id)
    if search:
        seats = seats.filter(
            Q(hall__name__icontains=search) |
            Q(row__icontains=search) |
            Q(number__icontains=search)
        )

    paginator = Paginator(seats, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    halls = Hall.objects.all().order_by('name')

    return render(request, 'ticket/admin_panel/seats.html', {
        'seats': page_obj,
        'halls': halls,
        'selected_hall': hall_id,
        'search': search,
    })


@require_POST
def admin_seat_delete(request, seat_id):
    """Удаление места (только если нет связанных билетов)"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    seat = get_object_or_404(Seat, id=seat_id)

    if Ticket.objects.filter(seat=seat).exists():
        messages.error(request, f'Нельзя удалить место "{seat}", так как на него есть билеты')
        return redirect('admin_panel_seats')

    seat_info = str(seat)
    seat.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='HALLS',
        description=f'Удалено место: {seat_info}',
        object_id=seat_id,
        object_repr=seat_info
    )
    messages.success(request, f'Место "{seat_info}" удалено')
    return redirect('admin_panel_seats')


# ==================== ЗАПРОСЫ ВОССТАНОВЛЕНИЯ ПАРОЛЯ ====================

@staff_member_required
def admin_password_reset_requests(request):
    """Просмотр запросов восстановления пароля"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    is_used = request.GET.get('is_used', '')

    requests_list = PasswordResetRequest.objects.select_related('user').all().order_by('-created_at')

    if search:
        requests_list = requests_list.filter(user__email__icontains=search)
    if is_used == 'yes':
        requests_list = requests_list.filter(is_used=True)
    elif is_used == 'no':
        requests_list = requests_list.filter(is_used=False)

    paginator = Paginator(requests_list, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/password_reset_requests.html', {
        'requests': page_obj,
        'search': search,
        'is_used_filter': is_used,
    })


@require_POST
def admin_password_reset_request_delete(request, pr_id):
    """Удаление запроса восстановления пароля"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    pr = get_object_or_404(PasswordResetRequest, id=pr_id)
    pr.delete()

    messages.success(request, 'Запрос удалён')
    return redirect('admin_panel_password_reset_requests')


# ==================== ЗАПРОСЫ СМЕНЫ EMAIL ====================

@staff_member_required
def admin_email_change_requests(request):
    """Просмотр запросов смены email"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    is_used = request.GET.get('is_used', '')

    requests_list = EmailChangeRequest.objects.select_related('user').all().order_by('-created_at')

    if search:
        requests_list = requests_list.filter(
            Q(user__email__icontains=search) | Q(new_email__icontains=search)
        )
    if is_used == 'yes':
        requests_list = requests_list.filter(is_used=True)
    elif is_used == 'no':
        requests_list = requests_list.filter(is_used=False)

    paginator = Paginator(requests_list, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/email_change_requests.html', {
        'requests': page_obj,
        'search': search,
        'is_used_filter': is_used,
    })


@require_POST
def admin_email_change_request_delete(request, ec_id):
    """Удаление запроса смены email"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    ec = get_object_or_404(EmailChangeRequest, id=ec_id)
    ec.delete()

    messages.success(request, 'Запрос удалён')
    return redirect('admin_panel_email_change_requests')


# ==================== API ТОКЕНЫ ====================

@staff_member_required
def admin_api_tokens(request):
    """Управление API токенами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    tokens = APIToken.objects.all().order_by('-is_active', 'label')

    return render(request, 'ticket/admin_panel/api_tokens.html', {'tokens': tokens})


@staff_member_required
def admin_api_token_add(request):
    """Добавление API токена"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    if request.method == 'POST':
        token = request.POST.get('token')
        label = request.POST.get('label')
        daily_limit = request.POST.get('daily_limit', 200)

        if APIToken.objects.filter(token=token).exists():
            messages.error(request, 'Токен с таким значением уже существует')
        else:
            api_token = APIToken.objects.create(
                token=token,
                label=label,
                daily_limit=daily_limit
            )
            OperationLogger.log_operation(
                request=request,
                action_type='CREATE',
                module_type='SYSTEM',
                description=f'Добавлен API токен: {label}',
                object_id=api_token.id,
                object_repr=label
            )
            messages.success(request, f'Токен "{label}" успешно добавлен')
            return redirect('admin_panel_api_tokens')

    return render(request, 'ticket/admin_panel/api_token_form.html', {'action': 'add'})


@staff_member_required
def admin_api_token_edit(request, token_id):
    """Редактирование API токена"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    token = get_object_or_404(APIToken, id=token_id)

    if request.method == 'POST':
        token.label = request.POST.get('label', token.label)
        token.daily_limit = request.POST.get('daily_limit', token.daily_limit)
        token.is_active = request.POST.get('is_active') == 'on'
        token.save()

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='SYSTEM',
            description=f'Обновлён API токен: {token.label}',
            object_id=token.id,
            object_repr=token.label
        )
        messages.success(request, f'Токен "{token.label}" обновлён')
        return redirect('admin_panel_api_tokens')

    return render(request, 'ticket/admin_panel/api_token_form.html', {
        'token': token,
        'action': 'edit'
    })


@require_POST
def admin_api_token_delete(request, token_id):
    """Удаление API токена"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    token = get_object_or_404(APIToken, id=token_id)
    label = token.label
    token.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SYSTEM',
        description=f'Удалён API токен: {label}',
        object_id=token_id,
        object_repr=label
    )
    messages.success(request, f'Токен "{label}" удалён')
    return redirect('admin_panel_api_tokens')


# ==================== КЭШ ИМПОРТА ====================

@staff_member_required
def admin_import_cache(request):
    """Просмотр кэша импорта"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    cache_type = request.GET.get('cache_type', '')
    caches = ImportCache.objects.all().order_by('-created_at')

    if cache_type:
        caches = caches.filter(cache_type=cache_type)

    paginator = Paginator(caches, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    return render(request, 'ticket/admin_panel/import_cache.html', {
        'caches': page_obj,
        'cache_type_filter': cache_type,
    })


@require_POST
def admin_import_cache_delete(request, cache_id):
    """Удаление записи кэша импорта"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    cache = get_object_or_404(ImportCache, id=cache_id)
    cache.delete()

    messages.success(request, 'Запись кэша удалена')
    return redirect('admin_panel_import_cache')


# ==================== ФИЛЬМЫ (полный CRUD) ====================

@staff_member_required
def admin_movies(request):
    """Управление фильмами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    search = request.GET.get('search', '')
    year = request.GET.get('year', '')

    movies = Movie.objects.select_related('age_rating').prefetch_related('genres', 'directors',
                                                                         'actors').all().order_by('-release_year',
                                                                                                  'title')

    if search:
        movies = movies.filter(title__icontains=search)
    if year:
        movies = movies.filter(release_year=year)

    paginator = Paginator(movies, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    years = Movie.objects.values_list('release_year', flat=True).distinct().order_by('-release_year')

    return render(request, 'ticket/admin_panel/admin_movies.html', {
        'movies': page_obj,
        'search': search,
        'selected_year': year,
        'years': years,
    })


@staff_member_required
def admin_movie_edit(request, movie_id):
    """Редактирование фильма"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    movie = get_object_or_404(Movie, id=movie_id)
    genres = Genre.objects.all().order_by('name')
    directors = Director.objects.all().order_by('surname', 'name')
    actors = Actor.objects.all().order_by('surname', 'name')
    countries = Country.objects.all().order_by('name')
    age_ratings = AgeRating.objects.all().order_by('name')

    if request.method == 'POST':
        movie.title = request.POST.get('title', movie.title)
        movie.short_description = request.POST.get('short_description', movie.short_description)
        movie.description = request.POST.get('description', movie.description)
        movie.duration = request.POST.get('duration', movie.duration)
        movie.release_year = request.POST.get('release_year', movie.release_year)
        movie.age_rating_id = request.POST.get('age_rating')

        if request.FILES.get('poster'):
            movie.poster = request.FILES['poster']

        movie.save()

        # Обновление связей
        movie.genres.set(request.POST.getlist('genres'))
        movie.directors.set(request.POST.getlist('directors'))
        movie.actors.set(request.POST.getlist('actors'))
        movie.countries.set(request.POST.getlist('countries'))

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='MOVIES',
            description=f'Обновлён фильм: {movie.title}',
            object_id=movie.id,
            object_repr=movie.title
        )
        messages.success(request, f'Фильм "{movie.title}" обновлён')
        return redirect('admin_panel_movies')

    return render(request, 'ticket/admin_panel/admin_movie_form.html', {
        'movie': movie,
        'genres': genres,
        'directors': directors,
        'actors': actors,
        'countries': countries,
        'age_ratings': age_ratings,
    })


@require_POST
def admin_movie_delete(request, movie_id):
    """Удаление фильма"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    movie = get_object_or_404(Movie, id=movie_id)
    title = movie.title

    # Проверка на связанные сеансы
    if movie.screenings.exists():
        messages.error(request, f'Нельзя удалить фильм "{title}", так как есть связанные сеансы')
        return redirect('admin_panel_movies')

    movie.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='MOVIES',
        description=f'Удалён фильм: {title}',
        object_id=movie_id,
        object_repr=title
    )
    messages.success(request, f'Фильм "{title}" удалён')
    return redirect('admin_panel_movies')


@staff_member_required
def admin_movie_add(request):
    """Добавление фильма"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    genres = Genre.objects.all().order_by('name')
    directors = Director.objects.all().order_by('surname', 'name')
    actors = Actor.objects.all().order_by('surname', 'name')
    countries = Country.objects.all().order_by('name')
    age_ratings = AgeRating.objects.all().order_by('name')

    if request.method == 'POST':
        title = request.POST.get('title')
        release_year = request.POST.get('release_year')
        duration = request.POST.get('duration')
        age_rating_id = request.POST.get('age_rating')
        short_description = request.POST.get('short_description', '')
        description = request.POST.get('description', '')

        movie = Movie.objects.create(
            title=title,
            release_year=release_year,
            duration=duration,
            age_rating_id=age_rating_id,
            short_description=short_description,
            description=description
        )

        movie.genres.set(request.POST.getlist('genres'))
        movie.directors.set(request.POST.getlist('directors'))
        movie.actors.set(request.POST.getlist('actors'))
        movie.countries.set(request.POST.getlist('countries'))

        if request.FILES.get('poster'):
            movie.poster = request.FILES['poster']
            movie.save()

        OperationLogger.log_operation(
            request=request,
            action_type='CREATE',
            module_type='MOVIES',
            description=f'Создан фильм: {title}',
            object_id=movie.id,
            object_repr=title
        )
        messages.success(request, f'Фильм "{title}" создан')
        return redirect('admin_panel_movies')

    return render(request, 'ticket/admin_panel/admin_movie_form.html', {
        'movie': None,
        'genres': genres,
        'directors': directors,
        'actors': actors,
        'countries': countries,
        'age_ratings': age_ratings,
    })


# ==================== СЕАНСЫ (полный CRUD) ====================

@staff_member_required
def admin_screenings(request):
    """Управление сеансами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    status = request.GET.get('status', 'upcoming')
    movie_id = request.GET.get('movie', '')
    hall_id = request.GET.get('hall', '')

    screenings = Screening.objects.select_related('movie', 'hall', 'hall__hall_type').all().order_by('start_time')

    if status == 'upcoming':
        screenings = screenings.filter(start_time__gt=timezone.now())
    elif status == 'past':
        screenings = screenings.filter(start_time__lt=timezone.now())

    if movie_id:
        screenings = screenings.filter(movie_id=movie_id)
    if hall_id:
        screenings = screenings.filter(hall_id=hall_id)

    paginator = Paginator(screenings, 30)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    movies = Movie.objects.all().order_by('title')
    halls = Hall.objects.all().order_by('name')

    return render(request, 'ticket/admin_panel/admin_screenings.html', {
        'screenings': page_obj,
        'movies': movies,
        'halls': halls,
        'status_filter': status,
        'selected_movie': movie_id,
        'selected_hall': hall_id,
    })


@staff_member_required
def admin_screening_edit(request, screening_id):
    """Редактирование сеанса"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    screening = get_object_or_404(Screening, id=screening_id)
    movies = Movie.objects.all().order_by('title')
    halls = Hall.objects.all().order_by('name')

    if request.method == 'POST':
        old_price = screening.ticket_price
        old_movie = screening.movie.title
        old_hall = screening.hall.name

        screening.movie_id = request.POST.get('movie')
        screening.hall_id = request.POST.get('hall')
        start_time_str = request.POST.get('start_time')

        if start_time_str:
            screening.start_time = timezone.make_aware(datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M'))

        # Сохраняем и пересчитываем цену
        screening.save()

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='SCREENINGS',
            description=f'Обновлён сеанс: {old_movie} → {screening.movie.title}',
            additional_data={
                'old_hall': old_hall,
                'new_hall': screening.hall.name,
                'old_price': float(old_price),
                'new_price': float(screening.ticket_price),
            }
        )
        messages.success(request, 'Сеанс обновлён')
        return redirect('admin_panel_screenings')

    return render(request, 'ticket/admin_panel/admin_screening_form.html', {
        'screening': screening,
        'movies': movies,
        'halls': halls,
    })


@require_POST
def admin_screening_delete(request, screening_id):
    """Удаление сеанса"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    screening = get_object_or_404(Screening, id=screening_id)

    if screening.tickets.exists():
        messages.error(request, f'Нельзя удалить сеанс, на который уже куплены билеты')
        return redirect('admin_panel_screenings')

    info = str(screening)
    screening.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='SCREENINGS',
        description=f'Удалён сеанс: {info}',
        object_id=screening_id,
        object_repr=info
    )
    messages.success(request, 'Сеанс удалён')
    return redirect('admin_panel_screenings')


# ==================== ЗАЛЫ (полный CRUD) ====================

@staff_member_required
def admin_halls(request):
    """Управление залами"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    halls = Hall.objects.select_related('hall_type').all().order_by('name')

    return render(request, 'ticket/admin_panel/admin_halls.html', {'halls': halls})


@staff_member_required
def admin_hall_edit(request, hall_id):
    """Редактирование зала"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    hall = get_object_or_404(Hall, id=hall_id)
    hall_types = HallType.objects.all().order_by('name')

    if request.method == 'POST':
        old_name = hall.name
        hall.name = request.POST.get('name', hall.name)
        hall.rows = request.POST.get('rows', hall.rows)
        hall.seats_per_row = request.POST.get('seats_per_row', hall.seats_per_row)
        hall.description = request.POST.get('description', hall.description)
        hall.hall_type_id = request.POST.get('hall_type')
        hall.save()

        OperationLogger.log_operation(
            request=request,
            action_type='UPDATE',
            module_type='HALLS',
            description=f'Обновлён зал: {old_name} → {hall.name}',
            object_id=hall.id,
            object_repr=hall.name
        )
        messages.success(request, f'Зал "{hall.name}" обновлён')
        return redirect('admin_panel_halls')

    return render(request, 'ticket/admin_panel/admin_hall_form.html', {
        'hall': hall,
        'hall_types': hall_types,
    })


@require_POST
def admin_hall_delete(request, hall_id):
    """Удаление зала"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    hall = get_object_or_404(Hall, id=hall_id)

    if hall.screenings.exists():
        messages.error(request, f'Нельзя удалить зал "{hall.name}", так как есть связанные сеансы')
        return redirect('admin_panel_halls')

    name = hall.name
    hall.delete()

    OperationLogger.log_operation(
        request=request,
        action_type='DELETE',
        module_type='HALLS',
        description=f'Удалён зал: {name}',
        object_id=hall_id,
        object_repr=name
    )
    messages.success(request, f'Зал "{name}" удалён')
    return redirect('admin_panel_halls')


# ==================== ДОБАВЛЕНИЕ СЕАНСА ====================

@staff_member_required
def admin_screening_add(request):
    """Добавление сеанса"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    movies = Movie.objects.all().order_by('title')
    halls = Hall.objects.all().order_by('name')

    if request.method == 'POST':
        movie_id = request.POST.get('movie')
        hall_id = request.POST.get('hall')
        start_time_str = request.POST.get('start_time')

        if not all([movie_id, hall_id, start_time_str]):
            messages.error(request, 'Заполните все обязательные поля')
        else:
            try:
                start_time = timezone.make_aware(datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M'))

                screening = Screening(
                    movie_id=movie_id,
                    hall_id=hall_id,
                    start_time=start_time
                )
                screening.save()  # Триггерит автоматический расчёт цены и времени окончания

                OperationLogger.log_operation(
                    request=request,
                    action_type='CREATE',
                    module_type='SCREENINGS',
                    description=f'Создан сеанс: {screening.movie.title} в {screening.hall.name}',
                    object_id=screening.id,
                    object_repr=str(screening)
                )
                messages.success(request, f'Сеанс успешно создан. Цена билета: {screening.ticket_price} ₽')
                return redirect('admin_panel_screenings')
            except Exception as e:
                messages.error(request, f'Ошибка при создании: {str(e)}')

    return render(request, 'ticket/admin_panel/admin_screening_form.html', {
        'screening': None,
        'movies': movies,
        'halls': halls,
    })


# ==================== ДОБАВЛЕНИЕ ЗАЛА ====================

@staff_member_required
def admin_hall_add(request):
    """Добавление зала"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    hall_types = HallType.objects.all().order_by('name')

    if request.method == 'POST':
        name = request.POST.get('name')
        hall_type_id = request.POST.get('hall_type')
        rows = request.POST.get('rows')
        seats_per_row = request.POST.get('seats_per_row')
        description = request.POST.get('description', '')

        if not all([name, hall_type_id, rows, seats_per_row]):
            messages.error(request, 'Заполните все обязательные поля')
        else:
            try:
                hall = Hall.objects.create(
                    name=name,
                    hall_type_id=hall_type_id,
                    rows=rows,
                    seats_per_row=seats_per_row,
                    description=description
                )
                OperationLogger.log_operation(
                    request=request,
                    action_type='CREATE',
                    module_type='HALLS',
                    description=f'Создан зал: {name}',
                    object_id=hall.id,
                    object_repr=name
                )
                messages.success(request, f'Зал "{name}" успешно создан. Создано {hall.total_seats} мест.')
                return redirect('admin_panel_halls')
            except Exception as e:
                messages.error(request, f'Ошибка при создании: {str(e)}')

    return render(request, 'ticket/admin_panel/admin_hall_form.html', {
        'hall': None,
        'hall_types': hall_types,
    })


@staff_member_required
def admin_reports(request):
    """Страница отчётов в админ-панели"""
    if not request.user.is_superuser:
        messages.error(request, 'У вас нет доступа к админ-панели.')
        return redirect('manager_dashboard')

    report_type = request.GET.get('report_type', 'revenue')
    period = request.GET.get('period', 'daily')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    report_data = None

    # Преобразуем строки в даты
    start_date_obj = None
    end_date_obj = None

    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        except ValueError:
            pass

    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
        except ValueError:
            pass

    # Генерируем отчёт в зависимости от типа
    if report_type == 'revenue':
        report_data = ReportGenerator.get_revenue_stats(
            period=period,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
    elif report_type == 'movies':
        report_data = ReportGenerator.get_popular_movies(
            limit=20,
            start_date=start_date_obj,
            end_date=end_date_obj
        )
    elif report_type == 'halls':
        report_data = ReportGenerator.get_hall_occupancy(
            start_date=start_date_obj,
            end_date=end_date_obj
        )
    elif report_type == 'sales':
        report_data = ReportGenerator.get_sales_statistics(
            start_date=start_date_obj,
            end_date=end_date_obj
        )

    # Обработка POST запроса для экспорта PDF
    if request.method == 'POST' and 'export_pdf' in request.POST:
        report_type = request.POST.get('report_type', 'revenue')
        period = request.POST.get('period', 'daily')
        start_date = request.POST.get('start_date', '')
        end_date = request.POST.get('end_date', '')

        # Снова получаем данные для экспорта
        start_date_obj = None
        end_date_obj = None

        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            except ValueError:
                pass

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                pass

        if report_type == 'revenue':
            export_data = ReportGenerator.get_revenue_stats(
                period=period,
                start_date=start_date_obj,
                end_date=end_date_obj
            )
        elif report_type == 'movies':
            export_data = ReportGenerator.get_popular_movies(
                limit=20,
                start_date=start_date_obj,
                end_date=end_date_obj
            )
        elif report_type == 'halls':
            export_data = ReportGenerator.get_hall_occupancy(
                start_date=start_date_obj,
                end_date=end_date_obj
            )
        elif report_type == 'sales':
            export_data = ReportGenerator.get_sales_statistics(
                start_date=start_date_obj,
                end_date=end_date_obj
            )
        else:
            export_data = None

        if export_data:
            # Название отчёта
            title_map = {
                'revenue': 'Выручка',
                'movies': 'Популярность фильмов',
                'halls': 'Загруженность залов',
                'sales': 'Продажи'
            }
            title = title_map.get(report_type, 'Отчёт')

            # Фильтры для PDF
            filters = {
                'period': period,
                'start_date': start_date_obj.strftime('%d.%m.%Y') if start_date_obj else None,
                'end_date': end_date_obj.strftime('%d.%m.%Y') if end_date_obj else None,
            }

            # Генерируем PDF с передачей пользователя
            pdf_buffer = generate_pdf_report(export_data, report_type, title, filters, user=request.user)

            # Логируем экспорт
            OperationLogger.log_operation(
                request=request,
                action_type='EXPORT',
                module_type='REPORTS',
                description=f'Экспорт отчёта "{title}" в PDF',
                additional_data={
                    'report_type': report_type,
                    'period': period,
                    'start_date': start_date,
                    'end_date': end_date,
                    'user': request.user.email
                }
            )

            # Отдаём PDF
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            filename = f"report_{report_type}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

    context = {
        'report_type': report_type,
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'report_data': report_data,
        'now': timezone.now(),
    }

    return render(request, 'ticket/admin_panel/reports.html', context)